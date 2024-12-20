import sys
import re
import csv
import sqlite3
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QLabel, QLineEdit, QPushButton, QTextEdit, QVBoxLayout, QFileDialog, QWidget, QMessageBox
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont
from docx import Document
from openpyxl import load_workbook

class URLExtractorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Link Extractor")
        self.setGeometry(100, 100, 600, 400)
        self.initUI()

    def initUI(self):
        font = QFont("Segoe UI", 11)

        # Central widget and layout
        self.central_widget = QWidget()
        self.layout = QVBoxLayout()
        self.layout.setSpacing(15)
        self.layout.setContentsMargins(20, 20, 20, 20)

        # Input CSV file selection
        self.input_label = QLabel("Select Input File:")
        self.input_label.setFont(font)
        self.layout.addWidget(self.input_label)

        self.input_field = QLineEdit()
        self.input_field.setFont(font)
        self.input_field.setPlaceholderText("Browse or enter the path to the input file...")
        self.layout.addWidget(self.input_field)

        self.input_button = QPushButton("Browse")
        self.input_button.setFont(font)
        self.input_button.clicked.connect(self.browse_input_file)
        self.layout.addWidget(self.input_button)

        # Output text file selection
        self.output_label = QLabel("Select Output Text File:")
        self.output_label.setFont(font)
        self.layout.addWidget(self.output_label)

        self.output_field = QLineEdit()
        self.output_field.setFont(font)
        self.output_field.setPlaceholderText("Browse or enter the path to the output text file...")
        self.layout.addWidget(self.output_field)

        self.output_button = QPushButton("Browse")
        self.output_button.setFont(font)
        self.output_button.clicked.connect(self.browse_output_file)
        self.layout.addWidget(self.output_button)

        # Extract button
        self.extract_button = QPushButton("Extract URLs")
        self.extract_button.setFont(font)
        self.extract_button.clicked.connect(self.extract_urls)
        self.extract_button.setStyleSheet("background-color: #4CAF50; color: white; padding: 10px; border-radius: 5px;")
        self.layout.addWidget(self.extract_button)

        # Output log
        self.log_label = QLabel("Log:")
        self.log_label.setFont(font)
        self.layout.addWidget(self.log_label)

        self.log_output = QTextEdit()
        self.log_output.setFont(font)
        self.log_output.setReadOnly(True)
        self.log_output.setStyleSheet("background-color: #555555; color: white; border: 1px solid #666; border-radius: 5px;")
        self.layout.addWidget(self.log_output)

        # Set layout
        self.central_widget.setLayout(self.layout)
        self.setCentralWidget(self.central_widget)

        # Styling
        self.setStyleSheet(""" 
            QMainWindow {
                background-color: #222222;
            }
            QLabel {
                color: #FFFFFF;
            }
            QPushButton {
                background-color: #2196F3;
                color: white;
                border-radius: 5px;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QLineEdit {
                background-color: #555555;
                color: #FFF;
                border: 0px solid #DDDDDD;
                border-radius: 5px;
                padding: 5px;
            }
            QMessageBox {
                background-color: #555555;
                color:black;
            }
        """)

    def browse_input_file(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Select Input File", "", "All Files (*);;CSV Files (*.csv);;Text Files (*.txt);;Word Files (*.docx);;Excel Files (*.xlsx);;SQLite Database (*.db)")
        if file_name:
            self.input_field.setText(file_name)

    def browse_output_file(self):
        file_name, _ = QFileDialog.getSaveFileName(self, "Select Output Text File", "", "Text Files (*.txt)")
        if file_name:
            self.output_field.setText(file_name)

    def extract_urls(self):
        input_file = self.input_field.text().strip()
        output_txt = self.output_field.text().strip()

        if not input_file or not output_txt:
            QMessageBox.critical(self, "Error", "Both input and output files must be specified.")
            return

        try:
            unique_urls = set()

            # Increase the CSV field size limit
            csv.field_size_limit(10**6)

            # Regular expression to match URLs with mandatory "://"
            url_pattern = re.compile(r'(http|ftp|https):\/\/([\w_-]+(?:\.[\w_-]+)+)(\/[\w.,@?^=%&:/~+#-]*)?')

            # Determine file type and extract URLs
            if input_file.endswith(".csv"):
                self.extract_from_csv(input_file, url_pattern, unique_urls)
            elif input_file.endswith(".txt"):
                self.extract_from_txt(input_file, url_pattern, unique_urls)
            elif input_file.endswith(".docx"):
                self.extract_from_docx(input_file, url_pattern, unique_urls)
            elif input_file.endswith(".xlsx"):
                self.extract_from_xlsx(input_file, url_pattern, unique_urls)
            elif input_file.endswith(".db") or input_file.endswith(".sqlite"):
                self.extract_from_db(input_file, url_pattern, unique_urls)

            # Write URLs to output text file
            with open(output_txt, 'w', encoding='utf-8') as txtfile:
                for url in sorted(unique_urls):
                    txtfile.write(url + '\n')

            self.log_output.append(f"Extracted {len(unique_urls)} unique URLs and saved to {output_txt}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred: {e}")

    def extract_from_csv(self, input_file, url_pattern, unique_urls):
        with open(input_file, 'r', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                for cell in row:
                    matches = url_pattern.findall(cell)
                    for match in matches:
                        scheme = match[0]
                        domain = match[1]
                        path = match[2] if match[2] else ''
                        full_url = f"{scheme}://{domain}{path}"
                        unique_urls.add(full_url)

    def extract_from_txt(self, input_file, url_pattern, unique_urls):
        with open(input_file, 'r', encoding='utf-8') as txtfile:
            content = txtfile.read()
            matches = url_pattern.findall(content)
            for match in matches:
                scheme = match[0]
                domain = match[1]
                path = match[2] if match[2] else ''
                full_url = f"{scheme}://{domain}{path}"
                unique_urls.add(full_url)

    def extract_from_docx(self, input_file, url_pattern, unique_urls):
        doc = Document(input_file)
        for para in doc.paragraphs:
            matches = url_pattern.findall(para.text)
            for match in matches:
                scheme = match[0]
                domain = match[1]
                path = match[2] if match[2] else ''
                full_url = f"{scheme}://{domain}{path}"
                unique_urls.add(full_url)

    def extract_from_xlsx(self, input_file, url_pattern, unique_urls):
        workbook = load_workbook(input_file, read_only=True)
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str):
                        matches = url_pattern.findall(cell)
                        for match in matches:
                            scheme = match[0]
                            domain = match[1]
                            path = match[2] if match[2] else ''
                            full_url = f"{scheme}://{domain}{path}"
                            unique_urls.add(full_url)

    def extract_from_db(self, input_file, url_pattern, unique_urls):
        try:
            conn = sqlite3.connect(input_file)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = cursor.fetchall()

            self.log_output.append(f"Tables in the database: {tables}")

            for table in tables:
                table_name = table[0]
                try:
                    cursor.execute(f"SELECT * FROM `{table_name}`")
                    rows = cursor.fetchall()
                    for row in rows:
                        for cell in row:
                            if isinstance(cell, str):  # Ensure the cell is a string
                                matches = url_pattern.findall(cell)
                                for match in matches:
                                    scheme = match[0]
                                    domain = match[1]
                                    path = match[2] if match[2] else ''
                                    full_url = f"{scheme}://{domain}{path}"
                                    unique_urls.add(full_url)
                except sqlite3.DatabaseError as e:
                    self.log_output.append(f"Error querying table {table_name}: {e}")
            conn.close()

        except sqlite3.DatabaseError as e:
            self.log_output.append(f"Error connecting to database: {e}")

# Main Application
if __name__ == "__main__":
    app = QApplication(sys.argv)

    window = URLExtractorApp()
    window.show()

    sys.exit(app.exec())
